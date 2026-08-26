#!/usr/bin/env python3
"""
PO Data Extractor
Extracts purchase order data from PDF files using Claude AI and exports to Excel.

Usage:
    python extract_pos.py --input ./pdfs --output po_data.xlsx
    python extract_pos.py --input ./pdfs --output po_data.xlsx --workers 5
    python extract_pos.py --input ./pdfs --output po_data.xlsx --force   # re-extract even if unchanged

Results are cached in a local SQLite database (po_data.db by default) keyed by
file content hash — rerunning over the same folder only re-extracts files that
are new or have changed. Excel output is generated from that data each run.

If the Anthropic account runs out of credits mid-run, the script pauses (exit
code 3) instead of burning retries — add credits and rerun the same command;
already-processed files are skipped automatically via the cache above.

Requirements:
    pip install -r requirements.txt
"""

import argparse
import base64
import io
import logging
import os
import re
import sys
import threading
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from itertools import permutations
from pathlib import Path

import pdfplumber
import anthropic

from math_check import validate_math
from price_check import flag_price_anomaly
from product_catalog import normalize_product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

import db


# ── Logging ─────────────────────────────────────────────────────────────────────
# INFO+ goes to a log file for after-the-fact debugging; only WARNING+ hits the
# console so it doesn't collide with the tqdm progress bar.

logger = logging.getLogger("extract_pos")


def configure_logging(log_path: str) -> None:
    logger.setLevel(logging.INFO)

    file_handler = logging.FileHandler(log_path)
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter(
        "%(asctime)s  %(levelname)-8s %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    ))
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.WARNING)
    console_handler.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
    logger.addHandler(console_handler)


# ── Configuration ──────────────────────────────────────────────────────────────

# The only valid products sold. Used for normalization after extraction.
VALID_PRODUCTS = ["Rainbow Mix", "Arugula", "Cilantro", "Bulls Blood Beets", "Genovese Basil", "Broccoli"]

# Valid container sizes in oz
VALID_SIZES = [1, 2, 3, 4, 8, 20]

EXTRACTION_PROMPT = """You are a data extraction assistant for Garfield Produce Company.
Extract all purchase order data from the document using the extract_po tool.

Rules:
- Extract every line item as a separate entry in line_items
- For dates, try to parse to YYYY-MM-DD format; if ambiguous leave as the original string
- Numbers should be numeric (not strings), strip currency symbols
- Preserve the full original product description in product_raw exactly as written
- sent_date: look for labels like 'Sent', 'Issued', 'Created', 'Date Sent', 'Transmitted'
- revision_number/revision_label: look for 'Rev', 'Revision', 'Revised', 'Amendment', 'Version', 'Ver'
- document_printed_at: some vendors print an UNLABELED date+time stamp in the page
  header (often top-right, near a 'Page X of Y' marker and/or the preparer's initials),
  e.g. '04/11/26 8:52p' — distinct from Issue Date / PO Date, which are day-only with
  no time. This is the exact moment this specific copy of the document was generated,
  and is often the ONLY signal (finer than any other date on the page) that
  distinguishes an updated reprint of the same PO number from the original — extract
  it exactly as printed (include the time) whenever present, even without an explicit
  label. Leave null if no such timestamp appears anywhere on the page.
- Some documents show a separate additional per-line charge on top of the main unit
  price — e.g. a column labeled 'Adtl. Cost', 'Additional Cost', 'Surcharge', or
  'Freight' that is added into that line's printed total. If a line has one, extract
  its total dollar amount for that line into additional_cost (per-unit rate × quantity,
  not the per-unit rate itself). If a line has no such charge, leave additional_cost
  null. line_total is always the full total exactly as printed, including any such
  charge — never subtract it out.
- Some inputs are a full email thread rather than a single document: each message is
  labeled by sender — "GARFIELD PRODUCE / US (email@address)" or "CUSTOMER
  (email@address)" — with its date, in chronological order, separated by "---".
  This represents a live negotiation, not a static document: quantities/items are
  often adjusted over the course of the conversation (e.g. due to availability), and
  a later message supersedes an earlier one for anything it changes. Read the WHOLE
  thread and extract the FINAL agreed order, not just the first message's ask. Only
  count what the CUSTOMER actually requested or confirmed as order content — a
  statement from GARFIELD PRODUCE / US (e.g. "we have Cilantro available today") is
  not itself a confirmed order line unless the customer accepts it. If the thread
  never actually results in a concrete order — e.g. it's pricing, logistics, or
  relationship correspondence with no items+quantities ever agreed — set is_po to
  false.
- Set is_po to false if the document is not a purchase order, and leave the other fields empty
"""

# Sent once as a cached system block (cache_control) instead of being resent
# in every user message — cuts repeated-prompt token cost across a batch run.
SYSTEM_BLOCK = [
    {
        "type": "text",
        "text": EXTRACTION_PROMPT,
        "cache_control": {"type": "ephemeral"},
    }
]

EXTRACTION_TOOL = {
    "name": "extract_po",
    "description": "Record structured purchase order data extracted from a document.",
    "input_schema": {
        "type": "object",
        "properties": {
            "is_po": {"type": "boolean", "description": "false if the document is not a purchase order"},
            "po_number": {"type": ["string", "null"]},
            "po_date": {"type": ["string", "null"]},
            "sent_date": {"type": ["string", "null"]},
            "delivery_date": {"type": ["string", "null"]},
            "document_printed_at": {
                "type": ["string", "null"],
                "description": "Unlabeled date+time page-header stamp exactly as printed "
                                "(e.g. '04/11/26 8:52p'), if present — see the rules above.",
            },
            "revision_number": {"type": ["string", "null"]},
            "revision_label": {"type": ["string", "null"]},
            "customer_name": {"type": ["string", "null"]},
            "customer_id": {"type": ["string", "null"]},
            "line_items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "product_raw": {"type": "string"},
                        "sku": {"type": ["string", "null"]},
                        "quantity": {"type": ["number", "null"]},
                        "unit_price": {"type": ["number", "null"]},
                        "additional_cost": {"type": ["number", "null"]},
                        "line_total": {"type": ["number", "null"]},
                    },
                    "required": ["product_raw"],
                },
            },
            "subtotal": {"type": ["number", "null"]},
            "tax": {"type": ["number", "null"]},
            "total": {"type": ["number", "null"]},
            "notes": {"type": ["string", "null"]},
        },
        "required": ["is_po", "line_items"],
    },
}

MODEL = "claude-opus-4-6"
MAX_RETRIES = 5          # was 3 — more attempts for transient rate limits
RETRY_DELAY = 10         # seconds base delay (exponential backoff from here)
MAX_TEXT_CHARS = 15000   # extracted PDF text longer than this is truncated before the API
                         # call (long multi-page POs); truncation is logged and flagged in
                         # the result's notes so a dropped trailing line item is visible
                         # instead of silent.


# ── Revision Detection & Diff Engine ──────────────────────────────────────────

_PRINTED_AT_PATTERN = re.compile(
    r"^\s*(\d{1,2})/(\d{1,2})/(\d{2,4})\s+(\d{1,2}):(\d{2})\s*([ap])m?\s*$", re.IGNORECASE
)


def _parse_printed_at(value):
    """Best-effort parse of document_printed_at (e.g. '04/11/26 8:52p') into a
    zero-padded 'YYYY-MM-DD HH:MM' string that sorts correctly alongside sent_date/
    po_date's YYYY-MM-DD format. Falls back to the raw value (still a usable, if not
    perfectly chronological, tiebreaker) if the format doesn't match what vendors
    have been observed to print."""
    if not value:
        return None
    m = _PRINTED_AT_PATTERN.match(value)
    if not m:
        return value
    month, day, year, hour, minute, ampm = m.groups()
    year = int(year)
    if year < 100:
        year += 2000
    hour = int(hour) % 12
    if ampm.lower() == "p":
        hour += 12
    try:
        return f"{year:04d}-{int(month):02d}-{int(day):02d} {hour:02d}:{minute}"
    except ValueError:
        return value


def _sort_key(result):
    """Sort key for ordering versions of the same PO: document_printed_at (the exact
    moment this copy was generated — the finest-grained and most reliable signal for
    same-day reprints/revisions, especially for customers who don't send an explicit
    revision confirmation) > source_received_at (the Gmail message's own timestamp,
    for POs/revisions ingested straight from email with no printed stamp of their
    own — already a zero-padded 'YYYY-MM-DD HH:MM:SS' string, no parsing needed) >
    sent_date > po_date > filename."""
    printed_at = _parse_printed_at(result.get("document_printed_at"))
    return (
        printed_at or result.get("source_received_at")
        or result.get("sent_date") or result.get("po_date") or "9999",
        result.get("_source_file") or ""
    )


def _item_key(item):
    """Unique identity of a line item: product + container size."""
    return (item.get("product_name", "UNKNOWN"), item.get("container_size", ""))


_DIFF_FIELDS = (
    ("quantity", "Qty", "{}"),
    ("unit_price", "Price", "${}"),
    ("additional_cost", "Adtl", "${}"),
    ("line_total", "Total", "${}"),
)


def _field_changes(p, c):
    """List of 'Label: old → new' strings for each tracked field that differs
    between two items of the same (product, size) key. Also doubles as the cost
    function _best_pairing uses (via len()) to score candidate correspondences."""
    changes = []
    for field, label, fmt in _DIFF_FIELDS:
        pv, cv = p.get(field), c.get(field)
        if pv != cv and not (pv is None and cv is None):
            changes.append(f"{label}: {fmt.format(pv)} → {fmt.format(cv)}")
    return changes


_MAX_EXHAUSTIVE_PAIRING = 6  # same-key duplicate groups (e.g. two "Rainbow Mix 1oz"
# lines at different prices) are matched by minimal total field-difference cost, not
# list position — positional pairing would flag a spurious "Changed" on both lines if
# the model simply listed them in a different order between revisions. Duplicate
# same-product+size lines are rare and almost always just 2, so an exhaustive
# permutation search is cheap; groups larger than this cap fall back to positional
# pairing to avoid factorial blowup on pathological input.


def _best_pairing(prev_items, curr_items):
    """Returns [(prev_item_or_None, curr_item_or_None), ...] for one (product,
    size) group, matching items by minimal total change cost (see _field_changes)
    rather than by list position — see _MAX_EXHAUSTIVE_PAIRING for the size cap
    and rationale."""
    n, m = len(prev_items), len(curr_items)
    if n == 0 or m == 0 or max(n, m) > _MAX_EXHAUSTIVE_PAIRING:
        return [
            (prev_items[i] if i < n else None, curr_items[i] if i < m else None)
            for i in range(max(n, m))
        ]

    # Search permutations of the shorter side's items into the longer side's slots
    # (only ever a handful of permutations at this group size).
    if n <= m:
        short, long_ = prev_items, curr_items
    else:
        short, long_ = curr_items, prev_items

    best_cost, best_assignment = None, None
    for perm in permutations(range(len(long_)), len(short)):
        cost = sum(len(_field_changes(short[i], long_[perm[i]])) for i in range(len(short)))
        if best_cost is None or cost < best_cost:
            best_cost, best_assignment = cost, perm

    matched = set(best_assignment)
    paired = [(short[i], long_[best_assignment[i]]) for i in range(len(short))]
    leftover = [(None, long_[j]) for j in range(len(long_)) if j not in matched]
    pairs = paired + leftover
    return pairs if n <= m else [(p, c) for c, p in pairs]


def _diff_items(prev_items, curr_items):
    """
    Compare two lists of line items. Returns (annotated, removed):
    - annotated: dict keyed by id(item) (for each item in curr_items) with a
      'revision_status' and 'changes' string.
    - removed: list of synthetic 'Removed' ghost entries for items dropped
      from prev_items.

    Items are matched by (product_name, container_size); since two real line
    items can legitimately share that key (e.g. the same product/size at two
    different prices on one PO), same-key items are matched via _best_pairing
    rather than collapsed to a single entry per key.

    Ghost 'Removed' rows injected by a previous diff (item.get('_removed'))
    are excluded from prev_items's grouping: they aren't real state from the
    previous version and must not be diffed again or they'd be re-flagged
    'Removed' and re-appended on every later revision.

    Keys are iterated in sorted order so the 'Removed' ghost rows this returns
    are in a deterministic order run-to-run (a plain set union's iteration
    order depends on Python's per-process string hash randomization).
    """
    prev_by_key = defaultdict(list)
    for i in prev_items:
        if i.get("_removed"):
            continue
        prev_by_key[_item_key(i)].append(i)

    curr_by_key = defaultdict(list)
    for i in curr_items:
        curr_by_key[_item_key(i)].append(i)

    annotated = {}
    removed = []

    for key in sorted(set(prev_by_key) | set(curr_by_key)):
        p_list = prev_by_key.get(key, [])
        c_list = curr_by_key.get(key, [])
        for p, c in _best_pairing(p_list, c_list):
            if c is None:
                removed_item = dict(p)
                removed_item["_removed"] = True
                removed_item["revision_status"] = "Removed"
                removed_item["changes"] = "Removed"
                removed.append(removed_item)
            elif p is None:
                annotated[id(c)] = ("Added", "Added")
            else:
                changes = _field_changes(p, c)
                if changes:
                    annotated[id(c)] = ("Changed", ", ".join(changes))
                else:
                    annotated[id(c)] = ("Unchanged", "")

    return annotated, removed


def annotate_revisions(all_results):
    """
    Groups results by PO number, sorts by date, assigns version labels,
    and diffs each version against its predecessor.
    Mutates each result and its line items in-place.
    """
    groups = defaultdict(list)

    for r in all_results:
        if "error" in r and "line_items" not in r:
            r["_version_label"] = ""
            r["_is_revision"] = False
            continue
        po_num = r.get("po_number") or r.get("_source_file")
        groups[po_num].append(r)

    for po_num, versions in groups.items():
        versions.sort(key=_sort_key)

        for v_idx, result in enumerate(versions):
            items = result.get("line_items") or []

            if v_idx == 0:
                # Original
                result["_version_label"] = "Original"
                result["_is_revision"] = False
                for item in items:
                    item["revision_status"] = "Original"
                    item["changes"] = ""
            else:
                # Revision
                rev_num = v_idx  # Rev 1, Rev 2, etc.
                explicit = result.get("revision_label") or result.get("revision_number")
                result["_version_label"] = explicit if explicit else f"Rev {rev_num}"
                result["_is_revision"] = True

                prev_items = versions[v_idx - 1].get("line_items") or []
                diff, removed_ghosts = _diff_items(prev_items, items)

                annotated_items = []
                for item in items:
                    status, changes = diff.get(id(item), ("Unchanged", ""))
                    item["revision_status"] = status
                    item["changes"] = changes if isinstance(changes, str) else ""
                    annotated_items.append(item)

                # Inject removed items as ghost rows
                annotated_items.extend(removed_ghosts)

                result["line_items"] = annotated_items


# ── PDF Text Extraction ────────────────────────────────────────────────────────

def extract_pdf_text(pdf_source: str | bytes) -> str | None:
    """Extract text from a PDF, given a file path or raw PDF bytes (e.g. a Gmail
    attachment fetched straight into memory, never touching disk). Returns None if
    extraction fails or yields no text."""
    try:
        source = io.BytesIO(pdf_source) if isinstance(pdf_source, bytes) else pdf_source
        with pdfplumber.open(source) as pdf:
            pages = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            full_text = "\n\n".join(pages).strip()
            return full_text if full_text else None
    except Exception:
        return None


def pdf_to_base64(pdf_source: str | bytes) -> str:
    """Convert a PDF (file path or raw bytes) to a base64 string for the API vision
    fallback."""
    if isinstance(pdf_source, bytes):
        return base64.standard_b64encode(pdf_source).decode("utf-8")
    with open(pdf_source, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8")


# ── Claude API Extraction ──────────────────────────────────────────────────────

def _is_out_of_credits(e: anthropic.PermissionDeniedError) -> bool:
    """True for Anthropic's 'credit balance too low' error, as opposed to other 403s."""
    if getattr(e, "type", None) == "billing_error":
        return True
    return "credit balance" in str(e).lower()


_NUMERIC_HEADER_FIELDS = ("subtotal", "tax", "total")
_NUMERIC_ITEM_FIELDS = ("quantity", "unit_price", "additional_cost", "line_total")


def _coerce_number(v):
    """Best-effort parse of a model-supplied numeric field. The extraction tool
    declares these as JSON numbers, but the model (especially the lighter
    text-thread model) occasionally returns them as strings — "12", "$8.50",
    "1,200.00" — which the tool schema does not coerce. Left alone, the first
    arithmetic that touches one raises (seen live: "unsupported operand type(s)
    for -: 'str' and 'str'" out of math_check), which the retry loop then
    mislabels a "Malformed tool response" and burns every attempt on.
    Returns a float, or None if missing / unparseable (treated as omitted)."""
    if v is None or isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        cleaned = v.strip().replace("$", "").replace(",", "").replace("%", "")
        if cleaned.lower() in ("", "-", "n/a", "na", "null", "none"):
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _coerce_numeric_fields(data: dict) -> None:
    """In-place coercion of every declared-numeric field in an extraction result,
    so the math/price checks and the values stored to Postgres are all real
    numbers regardless of how the model typed them."""
    for f in _NUMERIC_HEADER_FIELDS:
        if f in data:
            data[f] = _coerce_number(data[f])
    for item in data.get("line_items") or []:
        for f in _NUMERIC_ITEM_FIELDS:
            if f in item:
                item[f] = _coerce_number(item[f])


def _extract_from_source(
    client: anthropic.Anthropic,
    source_label: str,
    stop_event: threading.Event | None = None,
    reference_prices: dict | None = None,
    *,
    text: str | None = None,
    pdf_b64: str | None = None,
    extraction_method: str = "text",
    model: str = MODEL,
) -> dict | None:
    """
    Core Claude-calling extraction logic, decoupled from where the document content
    came from — a local PDF file (see extract_po_data, below) or content pulled
    straight from Gmail (attachment bytes or an email body's own text, never touching
    disk). Exactly one of text/pdf_b64 should be given; extraction_method records
    which, for logging and the extracted _extraction_method field.

    model defaults to MODEL (the local PDF pipeline's choice); the cloud Gmail
    pipeline passes its own — a whole email thread rendered to plain text is a much
    lighter extraction task than a scanned multi-page PO, and doesn't need the
    top-tier model the vision path benefits from.

    Returns None (instead of an error dict) if stop_event is already set when this
    call starts — used to fast-drain queued work after an out-of-credits pause without
    making further API calls or writing spurious error rows for files never attempted.

    reference_prices, if given, is a read-only {(customer, product, size): price} dict
    used to flag line items whose price looks anomalous — safe to share across worker
    threads since it's never mutated after being built.
    """
    text_truncated = bool(text) and len(text) > MAX_TEXT_CHARS
    if text_truncated:
        logger.warning(
            f"{source_label}: extracted text is {len(text)} chars, truncating to "
            f"{MAX_TEXT_CHARS} before sending to the model — trailing content may be lost"
        )
    last_error = "Unknown error"
    start_time = time.monotonic()

    for attempt in range(MAX_RETRIES):
        if stop_event is not None and stop_event.is_set():
            return None
        try:
            if text:
                messages = [
                    {
                        "role": "user",
                        "content": f"Document text:\n{text[:MAX_TEXT_CHARS]}"
                    }
                ]
            else:
                messages = [
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "document",
                                "source": {
                                    "type": "base64",
                                    "media_type": "application/pdf",
                                    "data": pdf_b64
                                }
                            },
                            {"type": "text", "text": "Extract the purchase order data from this document."}
                        ]
                    }
                ]

            response = client.messages.create(
                model=model,
                max_tokens=2000,
                system=SYSTEM_BLOCK,
                messages=messages,
                tools=[EXTRACTION_TOOL],
                tool_choice={"type": "tool", "name": "extract_po"}
            )

            tool_block = next((b for b in response.content if b.type == "tool_use"), None)
            if tool_block is None:
                raise ValueError("Model response did not include a tool_use block")

            data = dict(tool_block.input)
            is_po = data.pop("is_po", True)
            _coerce_numeric_fields(data)
            data["_source_file"] = source_label
            data["_extraction_method"] = extraction_method

            if not is_po:
                return {
                    "_source_file": source_label,
                    "_extraction_method": extraction_method,
                    "error": "not a purchase order"
                }

            # Normalize each line item's product name and container size
            for item in data.get("line_items") or []:
                product, size, is_sample, needs_review = normalize_product(
                    item.get("product_raw", ""),
                    unit_price=item.get("unit_price")
                )
                item["product_name"] = product
                item["container_size"] = size
                item["is_sample"] = is_sample
                item["needs_review"] = needs_review
                if reference_prices:
                    flag_price_anomaly(item, data.get("customer_name"), reference_prices)

            validate_math(data)

            if text_truncated:
                warning = (
                    f"⚠️ Source text was {len(text)} chars, truncated to {MAX_TEXT_CHARS} "
                    "before extraction — trailing content (e.g. line items) may be missing."
                )
                data["notes"] = f"{data['notes']}\n{warning}" if data.get("notes") else warning

            elapsed = time.monotonic() - start_time
            logger.info(
                f"{source_label}: extracted via {extraction_method} in {elapsed:.1f}s "
                f"(attempt {attempt + 1}/{MAX_RETRIES})"
            )
            return data

        except (ValueError, KeyError, TypeError) as e:
            last_error = f"Malformed tool response (attempt {attempt + 1}/{MAX_RETRIES}): {e}"
            logger.warning(f"{source_label}: {last_error}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
            continue

        except anthropic.PermissionDeniedError as e:
            if _is_out_of_credits(e):
                if stop_event is not None:
                    stop_event.set()
                logger.error(f"{source_label}: out of API credits — {e}")
                return {
                    "_source_file": source_label,
                    "_extraction_method": extraction_method,
                    "error": "insufficient API credits — add credits and rerun to resume"
                }
            # Other 403s (e.g. no access to this model) won't resolve by retrying.
            logger.warning(f"{source_label}: permission error — {e}")
            return {
                "_source_file": source_label,
                "_extraction_method": extraction_method,
                "error": f"Permission error: {e}"
            }

        except anthropic.RateLimitError as e:
            wait = RETRY_DELAY * (2 ** attempt)  # exponential: 5s, 10s, 20s
            last_error = f"Rate limit (attempt {attempt + 1}/{MAX_RETRIES}): {e} — retrying in {wait}s"
            logger.warning(f"{source_label}: rate limit hit, waiting {wait}s...")
            time.sleep(wait)
            continue

        except anthropic.APIStatusError as e:
            last_error = f"API error {e.status_code} (attempt {attempt + 1}/{MAX_RETRIES}): {e.message}"
            logger.warning(f"{source_label}: {last_error}")
            if attempt < MAX_RETRIES - 1:
                wait = RETRY_DELAY * (attempt + 1)
                time.sleep(wait)
            continue

        except anthropic.APIConnectionError as e:
            last_error = f"Connection error (attempt {attempt + 1}/{MAX_RETRIES}): {e}"
            logger.warning(f"{source_label}: {last_error}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
            continue

        except Exception as e:
            last_error = f"{type(e).__name__} (attempt {attempt + 1}/{MAX_RETRIES}): {e}"
            logger.warning(f"{source_label}: {last_error}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAY)
            continue

    logger.error(f"{source_label}: max retries exceeded — {last_error}")
    return {
        "_source_file": source_label,
        "_extraction_method": extraction_method,
        "error": f"Max retries exceeded — last error: {last_error}"
    }


def extract_po_data(
    client: anthropic.Anthropic,
    pdf_path: str,
    stop_event: threading.Event | None = None,
    reference_prices: dict | None = None,
) -> dict | None:
    """
    Extract PO data from a single local PDF file. Uses text extraction first, falls
    back to vision. Thin wrapper around _extract_from_source() that acquires content
    from a filesystem path — see that function for the shared Claude-calling logic
    (also used for Gmail-sourced content that never touches disk).
    """
    filename = os.path.basename(pdf_path)
    text = extract_pdf_text(pdf_path)
    pdf_b64 = None if text else pdf_to_base64(pdf_path)
    return _extract_from_source(
        client, filename, stop_event, reference_prices,
        text=text, pdf_b64=pdf_b64, extraction_method="text" if text else "vision",
    )


# ── Excel Output ───────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
BORDER      = Border(
    bottom=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="DDDDDD")
)
NORMAL_FONT = Font(name="Arial", size=10)


def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_data_cell(cell, row_idx):
    cell.font = NORMAL_FONT
    cell.border = BORDER
    cell.alignment = Alignment(vertical="center")
    if row_idx % 2 == 0:
        cell.fill = ALT_FILL


def build_excel(all_results: list[dict], output_path: str):
    wb = Workbook()

    # ── Sheet 1: All Line Items ────────────────────────────────────────────────
    ws_lines = wb.active
    ws_lines.title = "Line Items"

    line_headers = [
        "Source File", "PO Number", "Version", "PO Date", "Sent Date", "Delivery Date",
        "Customer Name", "Customer ID",
        "Product", "Container Size", "Sample", "Quantity",
        "Unit Price ($)", "Adtl. Cost ($)", "Line Total ($)", "PO Subtotal ($)", "PO Tax ($)", "PO Total ($)",
        "Revision Status", "Changes", "Math Check", "Price Anomaly",
        "SKU", "Notes", "Extraction Method"
    ]
    ws_lines.append(line_headers)
    style_header_row(ws_lines, 1, len(line_headers))
    ws_lines.row_dimensions[1].height = 30

    SAMPLE_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")  # green
    REVIEW_FILL   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")  # yellow
    REVISION_FILL = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")  # orange
    REMOVED_FILL  = PatternFill("solid", start_color="F2DCDB", end_color="F2DCDB")  # red-pink
    REMOVED_FONT  = Font(name="Arial", size=10, strikethrough=True, color="999999")

    data_row = 2
    for result in all_results:
        if "error" in result and "line_items" not in result:
            row = [
                result.get("_source_file", ""),
                f"ERROR: {result.get('error', 'unknown')}",
                *[""] * (len(line_headers) - 3),
                result.get("_extraction_method", "")
            ]
            ws_lines.append(row)
            for col in range(1, len(line_headers) + 1):
                cell = ws_lines.cell(row=data_row, column=col)
                cell.font = Font(name="Arial", size=10, color="CC0000")
            data_row += 1
            continue

        is_revision  = result.get("_is_revision", False)
        version_label = result.get("_version_label", "Original")

        items = result.get("line_items") or [{}]
        for item in items:
            product = item.get("product_name", "UNKNOWN")
            if product not in VALID_PRODUCTS and not item.get("_removed"):
                product = f"⚠️ {item.get('product_raw', 'UNKNOWN')}"

            is_sample     = item.get("is_sample", False)
            math_mismatch = item.get("math_mismatch", "")
            price_anomaly = item.get("price_anomaly", "")
            needs_review  = item.get("needs_review", False) or bool(math_mismatch) or bool(price_anomaly)
            rev_status    = item.get("revision_status", "")
            is_removed    = item.get("_removed", False)

            row = [
                result.get("_source_file", ""),
                result.get("po_number", ""),
                version_label,
                result.get("po_date", ""),
                result.get("sent_date", ""),
                result.get("delivery_date", ""),
                result.get("customer_name", ""),
                result.get("customer_id", ""),
                product,
                item.get("container_size", ""),
                "Yes" if is_sample else ("Review ⚠️" if needs_review else "No"),
                item.get("quantity"),
                item.get("unit_price"),
                item.get("additional_cost"),
                item.get("line_total"),
                result.get("subtotal"),
                result.get("tax"),
                result.get("total"),
                rev_status,
                item.get("changes", ""),
                math_mismatch,
                price_anomaly,
                item.get("sku", ""),
                result.get("notes", ""),
                result.get("_extraction_method", "")
            ]
            ws_lines.append(row)

            for col in range(1, len(line_headers) + 1):
                cell = ws_lines.cell(row=data_row, column=col)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
                if is_removed:
                    cell.fill = REMOVED_FILL
                    cell.font = REMOVED_FONT
                elif is_sample:
                    cell.fill = SAMPLE_FILL
                    cell.font = NORMAL_FONT
                elif needs_review:
                    cell.fill = REVIEW_FILL
                    cell.font = NORMAL_FONT
                elif is_revision:
                    cell.fill = REVISION_FILL
                    cell.font = NORMAL_FONT
                else:
                    style_data_cell(cell, data_row)
            data_row += 1

    # Column widths for Line Items
    col_widths = [22, 16, 10, 13, 13, 13, 26, 16, 18, 14, 9, 10, 14, 14, 14, 15, 11, 13, 14, 28, 24, 24, 14, 30, 16]
    for i, w in enumerate(col_widths, 1):
        ws_lines.column_dimensions[get_column_letter(i)].width = w

    ws_lines.freeze_panes = "A2"
    ws_lines.auto_filter.ref = ws_lines.dimensions

    # ── Sheet 2: PO Summary ────────────────────────────────────────────────────
    ws_po = wb.create_sheet("PO Summary")
    po_headers = [
        "Source File", "PO Number", "Version", "Is Revision",
        "PO Date", "Sent Date", "Delivery Date",
        "Customer Name", "Customer ID", "Line Item Count",
        "Subtotal ($)", "Tax ($)", "Total ($)", "Math Check", "Notes", "Extraction Method"
    ]
    ws_po.append(po_headers)
    style_header_row(ws_po, 1, len(po_headers))
    ws_po.row_dimensions[1].height = 30

    for i, result in enumerate(all_results, 2):
        items = result.get("line_items") or []
        is_rev = result.get("_is_revision", False)
        math_failed = result.get("math_check_failed", False)
        row = [
            result.get("_source_file", ""),
            result.get("po_number", ""),
            result.get("_version_label", ""),
            "Yes" if is_rev else "No",
            result.get("po_date", ""),
            result.get("sent_date", ""),
            result.get("delivery_date", ""),
            result.get("customer_name", ""),
            result.get("customer_id", ""),
            len(items),
            result.get("subtotal"),
            result.get("tax"),
            result.get("total"),
            result.get("math_check_detail", ""),
            result.get("notes", "") if "error" not in result else f"ERROR: {result.get('error')}",
            result.get("_extraction_method", "")
        ]
        ws_po.append(row)
        for col in range(1, len(po_headers) + 1):
            cell = ws_po.cell(row=i, column=col)
            if math_failed:
                cell.fill = REVIEW_FILL
                cell.font = NORMAL_FONT
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
            elif is_rev:
                cell.fill = REVISION_FILL
                cell.font = NORMAL_FONT
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
            else:
                style_data_cell(cell, i)

    po_widths = [22, 16, 10, 11, 13, 13, 13, 26, 16, 14, 14, 12, 14, 24, 30, 16]
    for i, w in enumerate(po_widths, 1):
        ws_po.column_dimensions[get_column_letter(i)].width = w
    ws_po.freeze_panes = "A2"
    ws_po.auto_filter.ref = ws_po.dimensions

    # ── Sheet 3: Errors Log ────────────────────────────────────────────────────
    errors = [r for r in all_results if "error" in r]
    if errors:
        ws_err = wb.create_sheet("Errors")
        ws_err.append(["Source File", "Error", "Extraction Method"])
        style_header_row(ws_err, 1, 3)
        for i, r in enumerate(errors, 2):
            ws_err.append([r.get("_source_file", ""), r.get("error", ""), r.get("_extraction_method", "")])
            for col in range(1, 4):
                style_data_cell(ws_err.cell(row=i, column=col), i)
        for col, w in zip("ABC", [28, 50, 18]):
            ws_err.column_dimensions[col].width = w

    wb.save(output_path)
    print(f"\n✅ Saved: {output_path}")
    print(f"   Sheets: Line Items, PO Summary{', Errors' if errors else ''}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Extract PO data from PDFs into Excel")
    parser.add_argument("--input",        required=True, help="Folder containing PDF files")
    parser.add_argument("--output",       default="po_data.xlsx", help="Output Excel file path")
    parser.add_argument("--workers",      type=int, default=3, help="Parallel workers (default: 3; reduce to 1 if hitting rate limits)")
    parser.add_argument("--limit",        type=int, default=None, help="Process only first N files (for testing)")
    parser.add_argument("--retry-failed", metavar="ERRORS_TXT", help="Retry only files listed in a text file (one filename per line)")
    parser.add_argument("--log-file",     default="extract_pos.log", help="Path to the log file (default: extract_pos.log)")
    parser.add_argument("--db",           default="po_data.db", help="SQLite database path (default: po_data.db)")
    parser.add_argument("--force",        action="store_true", help="Re-extract even if a file is unchanged since last run")
    args = parser.parse_args()

    configure_logging(args.log_file)
    logger.info(f"Run started: input={args.input} output={args.output} workers={args.workers}")

    input_dir = Path(args.input)
    if not input_dir.is_dir():
        print(f"❌ Error: '{args.input}' is not a directory", file=sys.stderr)
        logger.error(f"'{args.input}' is not a directory")
        sys.exit(1)

    all_pdfs = sorted(input_dir.glob("*.pdf")) + sorted(input_dir.glob("*.PDF"))
    if not all_pdfs:
        print(f"❌ No PDF files found in '{args.input}'", file=sys.stderr)
        logger.error(f"No PDF files found in '{args.input}'")
        sys.exit(1)

    # --retry-failed: filter to only the listed filenames
    if args.retry_failed:
        retry_path = Path(args.retry_failed)
        if not retry_path.exists():
            print(f"❌ Retry file not found: {args.retry_failed}", file=sys.stderr)
            logger.error(f"Retry file not found: {args.retry_failed}")
            sys.exit(1)
        retry_names = {line.strip() for line in retry_path.read_text().splitlines() if line.strip()}
        pdf_files = [p for p in all_pdfs if p.name in retry_names]
        print(f"🔁 Retry mode: {len(pdf_files)} file(s) from {args.retry_failed}")
        if not pdf_files:
            print("❌ None of the listed files found in input directory", file=sys.stderr)
            logger.error("None of the retry-listed files found in input directory")
            sys.exit(1)
    else:
        pdf_files = all_pdfs

    if args.limit:
        pdf_files = pdf_files[:args.limit]

    # Skip files whose content hasn't changed since a prior successful run
    conn = db.connect(args.db)
    db.init_db(conn)

    # Refresh reference prices from everything known *before* this run, so new
    # extractions are compared against the latest prior price, not a moving target.
    db.refresh_reference_prices(conn)
    reference_prices = db.get_reference_prices(conn)

    file_hashes = {}
    results = []
    to_extract = []
    for p in pdf_files:
        h = db.file_sha256(str(p))
        file_hashes[str(p)] = h
        cached = None if args.force else db.get_cached_result(conn, p.name, h)
        if cached is not None:
            results.append(cached)
        else:
            to_extract.append(p)

    if results:
        print(f"💾 {len(results)} file(s) unchanged since last run — reusing cached extraction")
    logger.info(f"{len(results)} cached, {len(to_extract)} to extract (force={args.force})")

    client = None
    if to_extract:
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            print("❌ ANTHROPIC_API_KEY environment variable not set", file=sys.stderr)
            logger.error("ANTHROPIC_API_KEY environment variable not set")
            sys.exit(1)
        client = anthropic.Anthropic(api_key=api_key)

    print(f"📄 Found {len(pdf_files)} PDF(s) in '{args.input}'  ({len(to_extract)} to extract)")
    print(f"🤖 Model: {MODEL}  |  Workers: {args.workers}  |  Max retries: {MAX_RETRIES}")
    print(f"📊 Output: {args.output}\n")

    errors = sum(1 for r in results if "error" in r)
    stop_event = threading.Event()

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(extract_po_data, client, str(p), stop_event, reference_prices): p
            for p in to_extract
        }
        with tqdm(total=len(futures), unit="pdf", ncols=80) as pbar:
            for future in as_completed(futures):
                pdf_path = futures[future]
                try:
                    result = future.result()
                except Exception as e:
                    result = {
                        "_source_file": pdf_path.name,
                        "_extraction_method": "unknown",
                        "error": f"{type(e).__name__}: {e}"
                    }
                if result is not None:
                    results.append(result)
                    db.save_result(conn, file_hashes[str(pdf_path)], result)
                    if "error" in result:
                        errors += 1
                    pbar.set_postfix({"errors": errors, "last": pdf_path.name[:20]})
                pbar.update(1)

    conn.close()

    if stop_event.is_set():
        remaining = len(pdf_files) - len(results)
        print(f"\n⏸️  Paused: ran out of API credits after processing {len(results)}/{len(pdf_files)} files.")
        print(f"   {remaining} file(s) not yet attempted — nothing was lost.")
        print(f"   👉 Add credits at https://console.anthropic.com/settings/billing, then rerun the same command:")
        print(f"      python3 {Path(__file__).name} --input {args.input} --output {args.output} "
              f"--db {args.db} --workers {args.workers}")
        print(f"   Already-processed files are cached in '{args.db}' and will be skipped automatically.")
        logger.error(f"Run paused: out of API credits after {len(results)}/{len(pdf_files)} files")

        results.sort(key=lambda r: (r.get("po_date") or "9999", r.get("_source_file") or ""))
        annotate_revisions(results)
        build_excel(results, args.output)
        sys.exit(3)

    # Sort results by PO date then filename for clean output
    results.sort(key=lambda r: (r.get("po_date") or "9999", r.get("_source_file") or ""))

    # Detect revisions and diff line items across versions of the same PO
    annotate_revisions(results)

    build_excel(results, args.output)

    success = len(results) - errors
    print(f"\n📈 Summary: {success} extracted successfully, {errors} errors")
    logger.info(f"Run finished: {success} succeeded, {errors} errors, output={args.output}")
    if errors:
        failed_names = [r["_source_file"] for r in results if "error" in r]
        failed_txt = Path(args.output).stem + "_failed.txt"
        Path(failed_txt).write_text("\n".join(failed_names))
        print(f"   ❌ Failed files saved to: {failed_txt}")
        print(f"   👉 Retry with: python3 {Path(__file__).name} --input {args.input} --output {args.output} --retry-failed {failed_txt} --workers 1")
        print(f"   Check the 'Errors' sheet in {args.output} for details on each failure")


if __name__ == "__main__":
    main()